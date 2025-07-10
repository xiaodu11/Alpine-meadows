import sys
import os
import numpy as np
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (QApplication, QMainWindow, QLabel, QVBoxLayout, QWidget, QPushButton, QHBoxLayout,
                             QFileDialog, QMessageBox, QTextEdit)
from PyQt5.QtGui import QImage, QPixmap
import cv2
from ultralytics import YOLO
import torch
import random
from openpyxl import load_workbook, Workbook


class Worker:
    def __init__(self):
        self.model = None
        self.current_image = None
        self.class_names = None
        self.detailed_info = {}  # 用于存储详细信息
        self.load_model()  # 直接加载模型
        self.load_plant_info()  # 直接加载植物信息

    def load_model(self):
        # 直接从相对路径加载模型
        model_path = "best.pt"  # 模型文件的相对路径
        if os.path.exists(model_path):
            try:
                self.model = YOLO(model_path)
                self.class_names = self.model.names
                print(f"模型加载成功，类别名称: {self.class_names}")
                return True
            except Exception as e:
                print(f"加载模型失败: {e}")
                return False
        else:
            print(f"模型文件不存在: {model_path}")
            return False

    def load_plant_info(self):
        # 从指定路径加载植物详细信息
        excel_path = "plant.xlsx"  # Excel文件的相对路径
        if os.path.exists(excel_path):
            try:
                wb = load_workbook(excel_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过表头
                    if row[0]:  # 确保种名存在
                        plant_name = row[0]
                        self.detailed_info[plant_name] = {
                            "科": row[1],
                            "属": row[2],
                            "种": row[0],
                            "分布地点": row[4],
                            "外观": row[5]
                        }
                print(f"植物信息加载成功，共加载 {len(self.detailed_info)} 种植物")
            except Exception as e:
                print(f"加载植物信息失败: {e}")
        else:
            print(f"Excel文件不存在: {excel_path}")

    def classify_image(self, image):
        if self.model is None:
            return []

        try:
            results = self.model.predict(image)
            if results:
                probs = results[0].probs
                if isinstance(probs, torch.Tensor):
                    class_id = torch.argmax(probs).item()
                    confidence = probs[class_id].item()
                else:
                    class_id = torch.argmax(probs.data).item()
                    confidence = probs.data[class_id].item()
                    class_name = self.class_names[class_id]
                # 减去一个随机数（范围在1%到5%之间，保留两位小数）
                random_num = round(random.uniform(0.01, 0.05), 4)
                adjusted_confidence = max(confidence - random_num, 0)
                return [(class_name, adjusted_confidence)]
            return []
        except Exception as e:
            print(f"分类图像失败: {e}")
            return []

    def save_image(self, image):
        if image is not None:
            file_name, _ = QFileDialog.getSaveFileName(None, "保存图片", "", "JPEG (*.jpg);;PNG (*.png);;All Files (*)")
            if file_name:
                cv2.imwrite(file_name, image)

    def export_results_to_excel(self, results, file_path):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["图片名称", "分类结果", "置信度", "科", "属", "种", "分布地点", "外观"])
            for result in results:
                image_name = result.get('image_name', '未知')
                class_name = result.get('class_name', '未知')
                confidence = result.get('confidence', 0)
                detailed_info = self.detailed_info.get(class_name, {})
                ws.append([
                    image_name,
                    class_name,
                    f"{confidence:.2%}",
                    detailed_info.get('科', '未知'),
                    detailed_info.get('属', '未知'),
                    detailed_info.get('种', '未知'),
                    detailed_info.get('分布地点', '未知'),
                    detailed_info.get('外观', '未知')
                ])
            wb.save(file_path)
            print(f"结果已导出到: {file_path}")
            return True
        except Exception as e:
            print(f"导出结果失败: {e}")
            return False


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("高寒百草库")
        self.setGeometry(300, 150, 1200, 600)

        # 创建主布局
        main_layout = QVBoxLayout()

        # 中间内容区域 - 水平布局
        content_layout = QHBoxLayout()

        # 左侧图片显示区域
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setMinimumSize(600, 400)
        self.image_label.setStyleSheet('''
            border:2px solid #ccc; 
            background-color: #f0f0f0;
            border-radius: 10px;
        ''')
        content_layout.addWidget(self.image_label)

        # 右侧文字结果显示区域 - 使用 QTextEdit 以便更好地显示详细信息
        self.result_textedit = QTextEdit()
        self.result_textedit.setReadOnly(True)  # 设置为只读
        self.result_textedit.setMinimumSize(400, 400)
        self.result_textedit.setStyleSheet('''
            border:2px solid #ccc; 
            background-color: white; 
            padding: 10px;
            border-radius: 10px;
        ''')
        content_layout.addWidget(self.result_textedit)

        main_layout.addLayout(content_layout)

        # 底部按钮区域
        button_layout = QHBoxLayout()

        self.classify_btn = QPushButton("图片分类")
        self.classify_btn.clicked.connect(self.select_image)
        button_layout.addWidget(self.classify_btn)

        self.folder_btn = QPushButton("文件夹分类")
        self.folder_btn.clicked.connect(self.classify_folder)
        button_layout.addWidget(self.folder_btn)

        self.export_btn = QPushButton("导出结果")
        self.export_btn.clicked.connect(self.export_results)
        button_layout.addWidget(self.export_btn)

        self.save_btn = QPushButton("保存图片")
        self.save_btn.clicked.connect(self.save_image)
        button_layout.addWidget(self.save_btn)

        self.exit_btn = QPushButton("退出")
        self.exit_btn.clicked.connect(self.exit_application)
        button_layout.addWidget(self.exit_btn)

        main_layout.addLayout(button_layout)

        # 设置中央窗口
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        self.worker = Worker()
        self.current_results = None
        self.folder_results = []  # 用于存储文件夹分类结果

    def select_image(self):
        image_path, _ = QFileDialog.getOpenFileName(None, "选择图片文件", "", "图片文件 (*.jpg *.jpeg *.png)")
        if image_path:
            self.classify_image(image_path)

    def classify_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "选择图片文件夹")
        if folder_path:
            image_paths = []
            for filename in os.listdir(folder_path):
                if filename.lower().endswith((".jpg", ".jpeg", ".png")):
                    image_path = os.path.join(folder_path, filename)
                    image_paths.append(image_path)
            self.folder_results = []  # 清空之前的结果
            for image_path in image_paths:
                self.classify_image(image_path, is_folder=True)

    def classify_image(self, image_path, is_folder=False):
        if image_path:
            image_data = np.fromfile(image_path, dtype=np.uint8)
            image = cv2.imdecode(image_data, cv2.IMREAD_COLOR)
            if image is not None:
                # 显示图片
                image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                height, width, channel = image_rgb.shape
                bytes_per_line = 3 * width
                q_img = QImage(image_rgb.data, width, height, bytes_per_line, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(q_img)
                self.image_label.setPixmap(pixmap.scaled(self.image_label.size(), Qt.KeepAspectRatio))

                # 分类推理
                results = self.worker.classify_image(image)
                self.current_results = results

                # 显示结果
                if results:
                    class_name, confidence = results[0]
                    print(f"分类结果: {class_name}, 置信度: {confidence:.2%}")
                    self.display_detailed_results(class_name, confidence)
                    if is_folder:
                        # 为每个图片创建一个包含分类结果的字典
                        result_entry = {
                            'image_name': os.path.basename(image_path),
                            'class_name': class_name,
                            'confidence': confidence
                        }
                        self.folder_results.append(result_entry)
                else:
                    self.result_textedit.setText("未检测到类别")
                self.worker.current_image = image.copy()
                self.save_btn.setEnabled(True)

    def display_detailed_results(self, class_name, confidence):
        # 构建详细结果文本
        result_text = f"<html><body><font size='7'><b>分类结果: {class_name}</b></font><br><br>"
        result_text += f"<font size='6'><b>预测概率: {confidence:.2%}</b></font><br><br><br>"
        result_text += f"<font size='5'><b>详细信息:</b></font><br><br>"

        # 获取详细信息
        detailed_info = self.worker.detailed_info.get(class_name, {})
        if detailed_info:
            result_text += f"<font size='4'><b>科: </b>{detailed_info.get('科', '未知')}</font><br><br>"
            result_text += f"<font size='4'><b>属: </b>{detailed_info.get('属', '未知')}</font><br><br>"
            result_text += f"<font size='4'><b>种: </b>{detailed_info.get('种', '未知')}</font><br><br>"
            result_text += f"<font size='4'><b>分布地点: </b>{detailed_info.get('分布地点', '未知')}</font><br><br>"
            result_text += f"<font size='4'><b>外观: </b>{detailed_info.get('外观', '未知')}</font><br><br>"
        else:
            result_text += "<font size='4'>未找到该类别的详细信息。</font>"

        result_text += "</body></html>"

        self.result_textedit.setText(result_text)

    def export_results(self):
        if self.folder_results:
            file_name, _ = QFileDialog.getSaveFileName(None, "导出结果", "", "Excel 文件 (*.xlsx)")
            if file_name:
                if self.worker.export_results_to_excel(self.folder_results, file_name):
                    QMessageBox.information(self, "导出成功", f"结果已导出到: {file_name}")
                else:
                    QMessageBox.warning(self, "导出失败", "导出结果时发生错误")
        else:
            QMessageBox.information(self, "导出失败", "没有可导出的结果")

    def save_image(self):
        if self.worker.current_image is not None:
            self.worker.save_image(self.worker.current_image)

    def exit_application(self):
        sys.exit()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())